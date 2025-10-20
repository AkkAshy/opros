import os
from datetime import datetime
from typing import Dict, List, Any
import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from config import EXPORTS_DIR
from database.db import DB_FILE

async def get_all_appeals() -> List[Dict[str, Any]]:
    """Получить все обращения с медиа"""
    async with aiosqlite.connect(DB_FILE) as db:
        cursor = await db.execute("""
            SELECT a.id, a.user_id, a.phone, a.full_name, a.address, a.domkom,
                   a.text, a.created_at, a.status, a.comment,
                   GROUP_CONCAT(m.file_path || ':' || m.file_type, ';') as media
            FROM appeals a
            LEFT JOIN media m ON a.id = m.appeal_id
            GROUP BY a.id
            ORDER BY a.created_at DESC
        """)
        rows = await cursor.fetchall()

        appeals = []
        for row in rows:
            appeal = {
                'id': row[0],
                'user_id': row[1],
                'phone': row[2],
                'full_name': row[3],
                'address': row[4],
                'domkom': row[5],
                'text': row[6],
                'created_at': row[7],
                'status': row[8],
                'comment': row[9] or '',
                'media_count': 0,
                'media_types': []
            }

            if row[10]:  # media
                media_items = row[10].split(';')
                appeal['media_count'] = len(media_items)
                appeal['media_types'] = list(set(item.split(':')[1] for item in media_items if ':' in item))

            appeals.append(appeal)

        return appeals

async def get_users_stats() -> Dict[str, Any]:
    """Получить статистику пользователей"""
    async with aiosqlite.connect(DB_FILE) as db:
        # Общее количество пользователей
        cursor = await db.execute("SELECT COUNT(*) FROM users")
        result = await cursor.fetchone()
        total_users = result[0] if result else 0

        # Количество админов
        cursor = await db.execute("SELECT COUNT(*) FROM users WHERE role = 'admin'")
        result = await cursor.fetchone()
        admin_count = result[0] if result else 0

        # Количество обычных пользователей
        user_count = total_users - admin_count

        return {
            'total_users': total_users,
            'admin_count': admin_count,
            'user_count': user_count
        }

async def get_appeals_stats() -> Dict[str, Any]:
    """Получить статистику обращений"""
    async with aiosqlite.connect(DB_FILE) as db:
        # Общее количество обращений
        cursor = await db.execute("SELECT COUNT(*) FROM appeals")
        result = await cursor.fetchone()
        total_appeals = result[0] if result else 0

        # Количество обработанных
        cursor = await db.execute("SELECT COUNT(*) FROM appeals WHERE status = 'processed'")
        result = await cursor.fetchone()
        processed_count = result[0] if result else 0

        # Количество необработанных
        unprocessed_count = total_appeals - processed_count

        # Статистика по медиа
        cursor = await db.execute("SELECT COUNT(*) FROM media")
        result = await cursor.fetchone()
        total_media = result[0] if result else 0

        cursor = await db.execute("SELECT file_type, COUNT(*) FROM media GROUP BY file_type")
        rows = await cursor.fetchall()
        media_types = {row[0]: row[1] for row in rows}

        return {
            'total_appeals': total_appeals,
            'processed_count': processed_count,
            'unprocessed_count': unprocessed_count,
            'total_media': total_media,
            'media_types': media_types
        }

async def create_excel_export_async() -> str:
    """Создать Excel файл с полной статистикой (асинхронная версия)"""
    # Получить данные асинхронно
    appeals = await get_all_appeals()
    users_stats = await get_users_stats()
    appeals_stats = await get_appeals_stats()

    # Создать workbook
    wb = Workbook()
    ws_stats = wb.active
    if ws_stats is None:
        raise ValueError("Не удалось создать лист статистики")
    ws_stats.title = "Статистика"

    # Заголовки
    headers = ["Ko'rsatkich", "Qiymat"]
    for col, header in enumerate(headers, 1):
        cell = ws_stats.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Данные статистики
    stats_data = [
        ["Jami foydalanuvchilar", users_stats['total_users']],
        ["Administratorlar", users_stats['admin_count']],
        ["Oddiy foydalanuvchilar", users_stats['user_count']],
        ["", ""],
        ["Jami murojaatlar", appeals_stats['total_appeals']],
        ["Ishlangan murojaatlar", appeals_stats['processed_count']],
        ["Ishlanmagan murojaatlar", appeals_stats['unprocessed_count']],
        ["", ""],
        ["Jami media fayllar", appeals_stats['total_media']],
    ]

    # Добавить типы медиа
    for media_type, count in appeals_stats['media_types'].items():
        media_type_uz = "Rasm" if media_type == "photo" else "Video" if media_type == "video" else media_type
        stats_data.append([f"{media_type_uz} turi", count])

    for row, (label, value) in enumerate(stats_data, 2):
        ws_stats.cell(row=row, column=1, value=label)
        ws_stats.cell(row=row, column=2, value=value)

    # Автоподгонка ширины колонок
    from openpyxl.utils import get_column_letter
    for col_num, column in enumerate(ws_stats.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            ws_stats.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # Лист обращений
    ws_appeals = wb.create_sheet("Murojaatlar")

    # Заголовки для обращений
    appeal_headers = ["ID", "Foydalanuvchi", "Telefon", "F.I.O.", "Manzil", "Uy MFI/OFI",
                     "Murojaat matni", "Yaratilgan sana", "Status", "Izoh",
                     "Media soni", "Media turlari"]

    for col, header in enumerate(appeal_headers, 1):
        cell = ws_appeals.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Данные обращений
    for row, appeal in enumerate(appeals, 2):
        ws_appeals.cell(row=row, column=1, value=appeal['id'])
        ws_appeals.cell(row=row, column=2, value=appeal['user_id'])
        ws_appeals.cell(row=row, column=3, value=appeal['phone'])
        ws_appeals.cell(row=row, column=4, value=appeal['full_name'])
        ws_appeals.cell(row=row, column=5, value=appeal['address'])
        ws_appeals.cell(row=row, column=6, value=appeal['domkom'])
        ws_appeals.cell(row=row, column=7, value=appeal['text'])
        ws_appeals.cell(row=row, column=8, value=appeal['created_at'])
        status_uz = "Ishlangan" if appeal['status'] == "processed" else "Ishlanmagan"
        ws_appeals.cell(row=row, column=9, value=status_uz)
        ws_appeals.cell(row=row, column=10, value=appeal['comment'])
        ws_appeals.cell(row=row, column=11, value=appeal['media_count'])
        media_types_uz = []
        for mt in appeal['media_types']:
            if mt == 'photo':
                media_types_uz.append('Rasm')
            elif mt == 'video':
                media_types_uz.append('Video')
            else:
                media_types_uz.append(mt)
        ws_appeals.cell(row=row, column=12, value=', '.join(media_types_uz))

    # Автоподгонка ширины колонок для обращений
    for col_num, column in enumerate(ws_appeals.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)  # Ограничить максимальную ширину
            ws_appeals.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # Сгенерировать имя файла с timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"statistics_{timestamp}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    # Сохранить файл
    wb.save(filepath)

    return filepath

def create_excel_export() -> str:
    """Создать Excel файл с полной статистикой (синхронная обертка)"""
    import asyncio
    return asyncio.run(create_excel_export_async())